import pandas as pd
from openpyxl import load_workbook

# File paths
file_paths = [
    r'C:\Users\unm30\Notebook\Github\d7047e\Project\BiLSTM-Multiple-Datasets\additional_datasets\OffenseEval\task_a_part1.xlsx',
    r'C:\Users\unm30\Notebook\Github\d7047e\Project\BiLSTM-Multiple-Datasets\additional_datasets\OffenseEval\task_a_part2.xlsx'
]
# Function to get the first sheet name
def get_first_sheet_name(file_path):
    workbook = load_workbook(filename=file_path, read_only=True)
    return workbook.sheetnames[0]

# Load and combine all Excel files
df_list = []
for file_path in file_paths:
    sheet_name = get_first_sheet_name(file_path)
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)  # Read without headers
    df.columns = ['ID', 'Message', 'Score1', 'Score2']  # Manually set column names
    df_list.append(df)

df_combined = pd.concat(df_list, ignore_index=True)

# Check if 'Message' column exists after cleaning
if 'Message' not in df_combined.columns:
    raise KeyError("The column 'Message' is not found in the combined data. Columns present: {}".format(df_combined.columns.tolist()))

# Clean the data by removing rows with NaN values in the 'Message' column
df_cleaned = df_combined.dropna(subset=['Message']).copy()  # Using .copy() to avoid SettingWithCopyWarning

# Define thresholds
threshold_high_score1 = 0.7
threshold_low_score1 = 0.2

# Function to assign clusters based on Score1 thresholds only
def assign_cluster(row):
    if row['Score1'] > threshold_high_score1:
        return 'High in Score1'
    elif row['Score1'] < threshold_low_score1:
        return 'Low in Score1'
    else:
        return 'Intermediate'

# Assign clusters
df_cleaned['CustomCluster'] = df_cleaned.apply(assign_cluster, axis=1)

# Filter out 'Intermediate' cluster
df_filtered = df_cleaned[df_cleaned['CustomCluster'] != 'Intermediate']

# Display the count of messages in each custom cluster
custom_cluster_summary = df_filtered['CustomCluster'].value_counts()
print("Custom Cluster Summary:\n", custom_cluster_summary)

# Function to save DataFrame to Excel in chunks
def save_to_excel_in_chunks(df, output_file_path, max_chunk_size=1048575):  # Ensuring chunk size is within Excel limit
    for i in range(0, len(df), max_chunk_size):
        chunk = df.iloc[i:i + max_chunk_size]
        chunk.to_excel(f"{output_file_path}_part{i // max_chunk_size + 1}.xlsx", index=False)
        print(f"Saved chunk {i // max_chunk_size + 1} to {output_file_path}_part{i // max_chunk_size + 1}.xlsx")

# Save each cluster to a separate Excel file in chunks
clusters = df_filtered['CustomCluster'].unique()
for cluster in clusters:
    cluster_df = df_filtered[df_filtered['CustomCluster'] == cluster]
    output_file_path = f"{cluster.replace(' ', '_')}_Cluster"
    save_to_excel_in_chunks(cluster_df, output_file_path)

# Display sample messages from each custom cluster
sample_custom_clusters = df_filtered.groupby('CustomCluster').apply(lambda x: x.sample(3, random_state=42)).reset_index(drop=True)
print("Sample Messages from Each Custom Cluster:\n", sample_custom_clusters[['CustomCluster', 'Message', 'Score1', 'Score2']])